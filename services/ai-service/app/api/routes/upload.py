import os
from typing import List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from sqlalchemy.orm import Session

from app.config import settings
from app.models.database import Audit, ChatSession, UploadedFile, get_db

from pypdf import PdfReader
try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

router = APIRouter(prefix="/upload", tags=["upload"])

os.makedirs(settings.upload_dir, exist_ok=True)


def _extract_text_from_file(path: str, content_type: str, filename: str) -> str:
    text = ""

    if content_type == "application/pdf" or filename.lower().endswith(".pdf"):
        reader = PdfReader(path)
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        text = "\n".join(parts)

    elif filename.lower().endswith(".docx") and DocxDocument:
        doc = DocxDocument(path)
        parts = [p.text for p in doc.paragraphs if p.text]
        text = "\n".join(parts)

    elif (filename.lower().endswith(".xlsx") or filename.lower().endswith(".xlsm")) and load_workbook:
        wb = load_workbook(path, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_vals = [str(v) for v in row if v is not None]
                if row_vals:
                    parts.append(" | ".join(row_vals))
        text = "\n".join(parts)

    else:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        except Exception:
            text = ""

    return text


@router.post("", status_code=status.HTTP_201_CREATED)
async def upload_file(
    audit_id: int = Form(...),
    session_id: str | None = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    audit = db.query(Audit).filter(Audit.id == audit_id).first()
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")

    # Session optional
    session = None
    if session_id:
        session = db.query(ChatSession).filter(ChatSession.id == session_id).first()
        if session and session.audit_id != audit.id:
            raise HTTPException(status_code=400, detail="Session gehört zu anderer Prüfung")

    if not session:
        session = ChatSession(audit_id=audit.id)
        db.add(session)
        db.commit()
        db.refresh(session)

    # Datei speichern
    audit_dir = os.path.join(settings.upload_dir, str(audit.id))
    os.makedirs(audit_dir, exist_ok=True)

    file_path = os.path.join(audit_dir, file.filename)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    # Text extrahieren
    extracted_text = _extract_text_from_file(
        path=file_path,
        content_type=file.content_type or "",
        filename=file.filename,
    )

    uploaded = UploadedFile(
        audit_id=audit.id,
        session_id=session.id,
        filename=file.filename,
        content_type=file.content_type,
        stored_path=file_path,
        extracted_text=extracted_text,
    )
    db.add(uploaded)
    db.commit()
    db.refresh(uploaded)

    return {
        "id": uploaded.id,
        "audit_id": uploaded.audit_id,
        "session_id": uploaded.session_id,
        "filename": uploaded.filename,
    }


@router.get("/audits/{audit_id}")
def list_files_for_audit(
    audit_id: int,
    db: Session = Depends(get_db),
):
    audit = db.query(Audit).filter(Audit.id == audit_id).first()
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")

    files: List[UploadedFile] = (
        db.query(UploadedFile)
        .filter(UploadedFile.audit_id == audit.id)
        .order_by(UploadedFile.created_at.desc())
        .all()
    )

    return [
        {
            "id": f.id,
            "filename": f.filename,
            "content_type": f.content_type,
            "created_at": f.created_at.isoformat() if f.created_at else None,
        }
        for f in files
    ]


@router.delete("/{file_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_file(
    file_id: str,
    db: Session = Depends(get_db),
):
    uploaded = db.query(UploadedFile).filter(UploadedFile.id == file_id).first()
    if not uploaded:
        raise HTTPException(status_code=404, detail="File not found")

    # Datei vom Filesystem löschen
    if uploaded.stored_path and os.path.exists(uploaded.stored_path):
        os.remove(uploaded.stored_path)

    db.delete(uploaded)
    db.commit()

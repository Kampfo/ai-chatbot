from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
import shutil
import os
import time
import logging
from pypdf import PdfReader
from sentence_transformers import SentenceTransformer
from vector_store import get_weaviate_client, init_schema, delete_by_filename
import asyncio
from concurrent.futures import ThreadPoolExecutor

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Document Service")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

model = None
executor = ThreadPoolExecutor(max_workers=1)

# Optional imports for DOCX/XLSX
try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None
    logger.warning("python-docx not installed, DOCX support disabled")

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None
    logger.warning("openpyxl not installed, XLSX support disabled")


def load_model():
    global model
    try:
        logger.info("Loading SentenceTransformer model...")
        model = SentenceTransformer('all-MiniLM-L6-v2')
        logger.info("Model loaded successfully.")
    except Exception as e:
        logger.error(f"Failed to load model: {e}")
        raise e


@app.on_event("startup")
async def startup_event():
    loop = asyncio.get_event_loop()
    try:
        await loop.run_in_executor(executor, load_model)
    except Exception as e:
        logger.error(f"Startup warning: Failed to load model: {e}")

    try:
        max_retries = 5
        for i in range(max_retries):
            try:
                logger.info(f"Attempting to connect to Weaviate (Attempt {i+1}/{max_retries})...")
                init_schema()
                logger.info("Weaviate schema initialized successfully.")
                break
            except Exception as e:
                logger.warning(f"Weaviate not ready yet: {e}")
                if i < max_retries - 1:
                    time.sleep(2)
    except Exception as e:
        logger.error(f"Startup warning: Weaviate init failed: {e}")


UPLOAD_DIR = "/app/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


@app.get("/health")
async def health():
    return {
        "status": "healthy",
        "model_loaded": model is not None,
    }


def _extract_text(file_path: str, filename: str) -> str:
    """Extract text from various file formats."""
    text = ""
    lower = filename.lower()

    if lower.endswith(".pdf"):
        try:
            reader = PdfReader(file_path)
            for page in reader.pages:
                text += (page.extract_text() or "") + "\n"
        except Exception as e:
            logger.error(f"Error reading PDF {filename}: {e}")

    elif lower.endswith(".docx") and DocxDocument:
        try:
            doc = DocxDocument(file_path)
            parts = [p.text for p in doc.paragraphs if p.text]
            text = "\n".join(parts)
        except Exception as e:
            logger.error(f"Error reading DOCX {filename}: {e}")

    elif (lower.endswith(".xlsx") or lower.endswith(".xlsm")) and load_workbook:
        try:
            wb = load_workbook(file_path, data_only=True)
            parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_vals = [str(v) for v in row if v is not None]
                    if row_vals:
                        parts.append(" | ".join(row_vals))
            text = "\n".join(parts)
        except Exception as e:
            logger.error(f"Error reading XLSX {filename}: {e}")

    else:
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        except Exception as e:
            logger.error(f"Error reading text file {filename}: {e}")

    return text


def process_file_sync(file_path: str, filename: str, audit_id: int):
    """Process file: extract text, chunk, embed, store in Weaviate."""
    try:
        logger.info(f"Processing file: {filename}")

        text_content = _extract_text(file_path, filename)

        if not text_content.strip():
            logger.warning(f"No text content in {filename}")
            return

        # Chunk by paragraphs, then by size
        chunk_size = 1000
        chunks = []
        paragraphs = text_content.split("\n\n")
        current_chunk = ""

        for para in paragraphs:
            if len(current_chunk) + len(para) > chunk_size and current_chunk:
                chunks.append(current_chunk.strip())
                current_chunk = para
            else:
                current_chunk += "\n\n" + para if current_chunk else para

        if current_chunk.strip():
            chunks.append(current_chunk.strip())

        if not chunks:
            chunks = [text_content[i:i+chunk_size] for i in range(0, len(text_content), chunk_size)]

        logger.info(f"Generated {len(chunks)} chunks for {filename}")

        client = get_weaviate_client()
        client.batch.configure(batch_size=100)

        with client.batch as batch:
            for chunk in chunks:
                embedding = model.encode(chunk).tolist()
                properties = {
                    "content": chunk,
                    "filename": filename,
                    "audit_id": audit_id,
                }
                batch.add_data_object(
                    data_object=properties,
                    class_name="Document",
                    vector=embedding,
                )

        logger.info(f"Successfully processed and uploaded {filename}")

    except Exception as e:
        logger.error(f"Failed to process file {filename}: {e}")


@app.post("/documents/upload")
async def upload_document(
    background_tasks: BackgroundTasks,
    audit_id: int = Form(...),
    file: UploadFile = File(...),
):
    if not model:
        raise HTTPException(status_code=503, detail="Model is still loading, please try again in a moment")

    try:
        file_location = f"{UPLOAD_DIR}/{file.filename}"
        with open(file_location, "wb+") as file_object:
            shutil.copyfileobj(file.file, file_object)

        background_tasks.add_task(process_file_sync, file_location, file.filename, audit_id)

        return {
            "filename": file.filename,
            "status": "processing_started",
            "message": "File uploaded and processing started in background",
        }

    except Exception as e:
        logger.error(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@app.post("/documents/search")
async def search_documents(query: str, audit_id: int = None):
    if not model:
        raise HTTPException(status_code=503, detail="Model not loaded")

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(executor, _search_sync, query, audit_id)


def _search_sync(query: str, audit_id: int):
    try:
        client = get_weaviate_client()
        query_vector = model.encode(query).tolist()

        where_filter = {}
        if audit_id:
            where_filter = {
                "path": ["audit_id"],
                "operator": "Equal",
                "valueInt": audit_id,
            }

        query_builder = client.query.get("Document", ["content", "filename", "audit_id"])
        query_builder = query_builder.with_near_vector({
            "vector": query_vector,
            "certainty": 0.6,
        })

        if audit_id:
            query_builder = query_builder.with_where(where_filter)

        return query_builder.with_limit(5).do()
    except Exception as e:
        logger.error(f"Search failed: {e}")
        raise e


@app.delete("/documents/{filename}")
async def delete_document_vectors(filename: str):
    """Delete all chunks for a given filename from Weaviate."""
    try:
        loop = asyncio.get_event_loop()
        deleted = await loop.run_in_executor(executor, delete_by_filename, filename)
        return {"filename": filename, "deleted": deleted}
    except Exception as e:
        logger.error(f"Delete failed: {e}")
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")
